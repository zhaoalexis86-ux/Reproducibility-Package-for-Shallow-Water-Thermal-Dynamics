import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ================= 路径配置区域（仅路径经过公开包适配） =================
REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
INPUT_FILE = REPOSITORY_ROOT / "results" / "reproduced" / "Figure6" / "Figure6_calculated_data_revised.csv"

OUTPUT_DIR = REPOSITORY_ROOT / "results" / "reproduced" / "Figure7"

OUTPUT_IMG = OUTPUT_DIR / "Figure7.png"

# 数值型结果表
OUTPUT_NUMERIC_CSV = OUTPUT_DIR / "Figure7_Table1_numeric_revised.csv"

# 论文展示用表格
OUTPUT_FORMATTED_CSV = OUTPUT_DIR / "Table1_source_v5_formatted_revised.csv"
OUTPUT_FORMATTED_XLSX = OUTPUT_DIR / "Table1_source_v5_formatted_revised.xlsx"

SEASONS = {
    "Spring": [3, 4, 5],
    "Summer": [6, 7, 8],
    "Autumn": [9, 10, 11],
    "Winter": [12, 1, 2]
}

SEASON_ORDER = ["Spring", "Summer", "Autumn", "Winter"]

# ================= 可直接微调的出版格式参数 =================
FIGURE_SIZE = (6.2, 3.8)       # 保持原图幅比例与尺寸
AXIS_LABEL_FONTSIZE = 15       # 在上一 revised 版基础上再放大 2 pt
TICK_LABEL_FONTSIZE = 12       # 在上一 revised 版基础上再放大 1 pt
LEGEND_FONTSIZE = 12           # 在上一 revised 版基础上再放大 1 pt
OUTPUT_DPI = 600

# ================= 字体设置 =================
plt.rcParams.update({
    "font.family": "sans-serif",
    "font.sans-serif": ["Arial", "DejaVu Sans"],
    "font.size": 12,
    "axes.linewidth": 0.9,
    "axes.labelsize": AXIS_LABEL_FONTSIZE,
    "axes.titlesize": 12,
    "xtick.labelsize": TICK_LABEL_FONTSIZE,
    "ytick.labelsize": TICK_LABEL_FONTSIZE,
    "legend.fontsize": LEGEND_FONTSIZE,
    "xtick.direction": "in",
    "ytick.direction": "in",
    "mathtext.fontset": "stix",
    "axes.grid": False,
    "axes.unicode_minus": False
})

# ================= 配色 =================
COLOR_RED = "#E76354"
COLOR_YELLOW = "#F7D06A"
COLOR_LIGHT_BLUE = "#72BCD4"
COLOR_DEEP_BLUE = "#376B95"

COLOR_MAP = {
    r"$Q_{sw}$": COLOR_RED,
    r"$Q_{lw}$": COLOR_YELLOW,
    r"$Q_e$": COLOR_LIGHT_BLUE,
    r"$Q_h$": COLOR_DEEP_BLUE
}


def read_input_file(path):
    path = Path(path)

    if path.suffix.lower() == ".csv":
        try:
            df = pd.read_csv(path, index_col=0, parse_dates=True, encoding="utf-8-sig")
        except UnicodeDecodeError:
            df = pd.read_csv(path, index_col=0, parse_dates=True, encoding="gb18030")
    elif path.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(path, index_col=0, parse_dates=True)
    else:
        raise ValueError("仅支持 csv / xlsx / xls 文件。")

    df.index = pd.to_datetime(df.index, errors="coerce")
    df = df[~df.index.isna()].sort_index()

    return df


def find_flux_columns(df):
    """
    兼容不同版本列名。
    前一段热通量代码输出列名为：
    Q_sw, Q_lw, Q_e, Q_h, Q_net
    """
    candidate_cols = {
        r"$Q_{sw}$": ["Q_sn(Solar)", "Q_sw", "Qsw", "Q_sn"],
        r"$Q_{lw}$": ["Q_ln(Longwave)", "Q_lw", "Qlw", "Q_ln"],
        r"$Q_e$": ["Q_e(Latent)", "Q_e", "Qe"],
        r"$Q_h$": ["Q_h(Sensible)", "Q_h", "Qh"]
    }

    flux_cols = {}

    for label, candidates in candidate_cols.items():
        found = None

        for c in candidates:
            if c in df.columns:
                found = c
                break

        if found is None:
            raise KeyError(f"未找到 {label} 对应列，候选列为：{candidates}")

        flux_cols[label] = found

    return flux_cols


def find_qnet_column(df):
    """
    查找 Qnet 列。如果没有，则后续用四项通量代数和计算。
    """
    candidates = ["Q_net", "Qnet", "Q_net(W/m2)", "Qnet(W/m2)", "Q_net(Net)", "Qnet(Net)"]

    for c in candidates:
        if c in df.columns:
            return c

    return None


def calc_stats_table(df, flux_cols, qnet_col=None):
    """
    输出两类信息：
    1. 各通量的季节平均代数值 mean_signed，单位 W m-2；
    2. 各通量的季节贡献率 contribution，基于 mean(abs(Qi)) / sum(mean(abs(Qi)))。
    """

    rows = []

    # ================= 分季节统计 =================
    for season_name in SEASON_ORDER:
        months = SEASONS[season_name]
        df_season = df[df.index.month.isin(months)].copy()

        # 去除四项通量任一缺失的时刻
        df_season = df_season.dropna(subset=list(flux_cols.values()))

        if df_season.empty:
            continue

        row = {
            "Season": season_name,
            "Days": df_season.index.normalize().nunique()
        }

        signed_means = {}
        abs_means = {}

        for label, col in flux_cols.items():
            signed_means[label] = df_season[col].mean()
            abs_means[label] = df_season[col].abs().mean()

        total_abs_activity = sum(abs_means.values())

        for label in flux_cols.keys():
            row[f"{label}_mean"] = signed_means[label]
            row[f"{label}_contribution"] = (
                abs_means[label] / total_abs_activity * 100
                if total_abs_activity > 0 else np.nan
            )

        # Qnet
        if qnet_col is not None and qnet_col in df_season.columns:
            row[r"$Q_{net}$"] = df_season[qnet_col].mean()
        else:
            row[r"$Q_{net}$"] = sum(signed_means.values())

        rows.append(row)

    # ================= 全年平均统计 =================
    df_year = df.dropna(subset=list(flux_cols.values())).copy()

    annual_row = {
        "Season": "Annual average",
        "Days": df_year.index.normalize().nunique()
    }

    signed_means = {}
    abs_means = {}

    for label, col in flux_cols.items():
        signed_means[label] = df_year[col].mean()
        abs_means[label] = df_year[col].abs().mean()

    total_abs_activity = sum(abs_means.values())

    for label in flux_cols.keys():
        annual_row[f"{label}_mean"] = signed_means[label]
        annual_row[f"{label}_contribution"] = (
            abs_means[label] / total_abs_activity * 100
            if total_abs_activity > 0 else np.nan
        )

    if qnet_col is not None and qnet_col in df_year.columns:
        annual_row[r"$Q_{net}$"] = df_year[qnet_col].mean()
    else:
        annual_row[r"$Q_{net}$"] = sum(signed_means.values())

    rows.append(annual_row)

    df_stats = pd.DataFrame(rows)

    return df_stats


def format_flux_cell(value, contribution):
    """
    生成论文表格中的形式：
    +139.2
    (72.0%)
    """
    if pd.isna(value) or pd.isna(contribution):
        return ""

    sign = "+" if value >= 0 else "-"
    return f"{sign}{abs(value):.1f}\n({contribution:.1f}%)"


def make_formatted_table(df_stats):
    """
    生成论文展示用表格。
    """

    flux_labels = [r"$Q_{sw}$", r"$Q_{lw}$", r"$Q_e$", r"$Q_h$"]

    formatted_rows = []

    for _, row in df_stats.iterrows():
        out = {
            "": row["Season"],
            "Days": int(row["Days"]) if pd.notna(row["Days"]) else ""
        }

        for label in flux_labels:
            out[label] = format_flux_cell(
                row[f"{label}_mean"],
                row[f"{label}_contribution"]
            )

        qnet_val = row[r"$Q_{net}$"]
        out[r"$Q_{net}$"] = f"{qnet_val:.1f}" if pd.notna(qnet_val) else ""

        formatted_rows.append(out)

    df_formatted = pd.DataFrame(formatted_rows)

    return df_formatted


def save_formatted_excel(df_formatted, output_xlsx):
    """
    输出接近论文表格样式的 Excel。
    """
    df_formatted.to_excel(output_xlsx, index=False)

    wb = load_workbook(output_xlsx)
    ws = wb.active
    ws.title = "Seasonal flux table"

    green_fill = PatternFill("solid", fgColor="C6E0B4")
    thin_side = Side(style="thin", color="000000")
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.fill = green_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.font = Font(name="Times New Roman", size=11)
            cell.border = border

    for cell in ws[1]:
        cell.font = Font(name="Times New Roman", size=11, bold=True)

    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 34

    col_widths = {
        "A": 16,
        "B": 10,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 12
    }

    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_xlsx)


def plot_contribution_bars(df_stats):
    print(">>> 正在绘制季节贡献率柱状图...")

    # 只绘制四个季节，不绘制 Annual average
    df_plot = df_stats[df_stats["Season"].isin(SEASON_ORDER)].copy()
    df_plot = df_plot.set_index("Season").loc[SEASON_ORDER]

    order = [r"$Q_{sw}$", r"$Q_{lw}$", r"$Q_e$", r"$Q_h$"]

    fig, ax = plt.subplots(figsize=FIGURE_SIZE)

    n_seasons = len(df_plot.index)
    n_vars = len(order)
    bar_width = 0.18
    indices = np.arange(n_seasons)

    for i, label in enumerate(order):
        offset = (i - (n_vars - 1) / 2) * bar_width

        ax.bar(
            indices + offset,
            df_plot[f"{label}_contribution"],
            width=bar_width,
            label=label,
            color=COLOR_MAP[label],
            edgecolor="black",
            linewidth=0.6
        )

    ax.set_ylabel("Contribution (%)")
    ax.set_xticks(indices)
    ax.set_xticklabels(df_plot.index)
    ax.set_ylim(0, 80)
    ax.yaxis.set_major_locator(ticker.MultipleLocator(10))

    ax.legend(
        loc="upper center",
        bbox_to_anchor=(0.67, 1.00),
        ncol=2,
        frameon=False,
        handlelength=1.6,
        columnspacing=1.0,
        borderpad=0.2
    )

    ax.spines["top"].set_visible(True)
    ax.spines["right"].set_visible(True)
    ax.grid(False)

    plt.tight_layout()
    plt.savefig(OUTPUT_IMG, dpi=OUTPUT_DPI, bbox_inches="tight")
    fig.savefig(OUTPUT_DIR / "Figure7.eps", format="eps", bbox_inches="tight")
print(f"[OK] 图片已保存至: {OUTPUT_IMG}")


def step3_contribution_analysis():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(">>> 正在进行季节性热通量幅值与贡献率分析...")

    df = read_input_file(INPUT_FILE)

    flux_cols = find_flux_columns(df)
    qnet_col = find_qnet_column(df)

    print(">>> 识别到的通量列：")
    for label, col in flux_cols.items():
        print(f"   {label}: {col}")

    if qnet_col is not None:
        print(f"   Qnet: {qnet_col}")
    else:
        print("   未找到 Qnet 列，将使用四项通量代数和计算 Qnet。")

    df_stats = calc_stats_table(df, flux_cols, qnet_col=qnet_col)

    # 保存数值型结果表
    df_stats_round = df_stats.copy()

    for col in df_stats_round.columns:
        if col not in ["Season", "Days"]:
            df_stats_round[col] = df_stats_round[col].round(2)

    df_stats_round.to_csv(OUTPUT_NUMERIC_CSV, index=False, encoding="utf-8-sig")
    print(f"[OK] 数值型结果表已保存: {OUTPUT_NUMERIC_CSV}")

    # 生成论文展示表
    df_formatted = make_formatted_table(df_stats)

    df_formatted.to_csv(OUTPUT_FORMATTED_CSV, index=False, encoding="utf-8-sig")
    print(f"[OK] 论文展示用 CSV 表格已保存: {OUTPUT_FORMATTED_CSV}")

    save_formatted_excel(df_formatted, OUTPUT_FORMATTED_XLSX)
    print(f"[OK] 论文展示用 Excel 表格已保存: {OUTPUT_FORMATTED_XLSX}")

    print(">>> 论文展示表预览：")
    print(df_formatted)

    plot_contribution_bars(df_stats)


if __name__ == "__main__":
    step3_contribution_analysis()
